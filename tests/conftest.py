"""Pytest configuration and fixtures for bcra-pagos-minoristas tests."""

import tempfile
from datetime import datetime
from pathlib import Path
from typing import Generator

import openpyxl
import pandas as pd
import pytest

from bcra_pagos_minoristas_xlsx_to_parquet.config import AppConfig
from bcra_pagos_minoristas_xlsx_to_parquet.models import (
    DiscoveredFile,
    DiscoveryResult,
)


@pytest.fixture
def temp_dir() -> Generator[Path, None, None]:
    """Provide a temporary directory for test outputs."""
    with tempfile.TemporaryDirectory() as tmpdir:
        yield Path(tmpdir)


@pytest.fixture
def config(temp_dir: Path) -> AppConfig:
    """Provide a test configuration."""
    return AppConfig()


@pytest.fixture
def sample_simple_excel(temp_dir: Path) -> Path:
    """Create a minimal sample XLSX file for testing."""
    file_path = temp_dir / "sample_simple.xlsx"

    # Create a simple DataFrame with basic structure
    data = {
        "Concepto": ["Cheques", "Transferencias"],
        "Cantidad": [100, 200],
        "Monto": [10000, 20000],
    }
    df = pd.DataFrame(data)

    # Write to Excel
    df.to_excel(file_path, index=False, engine="openpyxl")

    return file_path


@pytest.fixture
def sample_multirow_header_excel(temp_dir: Path) -> Path:
    """Create a sample XLSX file with multi-row headers (spec P3)."""
    file_path = temp_dir / "sample_multirow.xlsx"

    # Create workbook manually with multi-row headers
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Datos"

    # Row 1: Main header
    ws["A1"] = "Movimiento"
    ws["B1"] = "Movimiento"
    ws["C1"] = "Monto"

    # Row 2: Sub-header
    ws["A2"] = "Entidad"
    ws["B2"] = "Canal"
    ws["C2"] = "Total"

    # Row 3: Data
    ws["A3"] = "Banco A"
    ws["B3"] = "ATM"
    ws["C3"] = 5000

    ws["A4"] = "Banco B"
    ws["B4"] = "WEB"
    ws["C4"] = 7500

    wb.save(file_path)

    return file_path


@pytest.fixture
def sample_discovery_html(temp_dir: Path) -> Path:
    """Create a sample HTML file simulating BCRA discovery page (spec R1)."""
    file_path = temp_dir / "discovery_sample.html"

    html_content = """
    <html>
    <body>
        <a href="https://www.bcra.gob.ar/Estadisticas/datos/2024-05-Cheques.xlsx">
            Mayo 2024 - Cheques
        </a>
        <a href="https://www.bcra.gob.ar/Estadisticas/datos/2024-04-Cheques.xlsx">
            Abril 2024 - Cheques
        </a>
        <a href="https://www.bcra.gob.ar/Estadisticas/datos/2024-05-Transferencias.xlsx">
            Mayo 2024 - Transferencias
        </a>
    </body>
    </html>
    """

    file_path.write_text(html_content)
    return file_path


@pytest.fixture
def sample_normalized_data() -> pd.DataFrame:
    """Provide normalized sample data (spec N1-N9)."""
    return pd.DataFrame(
        {
            "concepto": ["cheques", "transferencias", "tarjetas"],
            "cantidad": [100, 200, 150],
            "monto_total": [10000.50, 25000.75, 15000.25],
            "periodo": ["2024-05-01", "2024-05-01", "2024-05-01"],
            "period_boundary": ["month_end", "month_end", "month_end"],
            "ingested_at": [
                "2024-05-25T10:00:00Z",
                "2024-05-25T10:00:00Z",
                "2024-05-25T10:00:00Z",
            ],
        }
    )


@pytest.fixture
def mock_discovery_result() -> DiscoveryResult:
    """Provide a mock discovery result (spec R7)."""
    selected_file = DiscoveredFile(
        url="https://www.bcra.gob.ar/Estadisticas/datos/2024-05-Cheques.xlsx",
        filename="2024-05-Cheques.xlsx",
        version="2024-05",
        modified_at=datetime(2024, 5, 25, 10, 0, 0),
    )

    return DiscoveryResult(
        source_url="https://www.bcra.gob.ar",
        selected=selected_file,
        candidates=[selected_file],
        discovered_at=datetime(2024, 5, 25, 10, 0, 0),
        selection_reason="Latest version matching pattern",
    )
